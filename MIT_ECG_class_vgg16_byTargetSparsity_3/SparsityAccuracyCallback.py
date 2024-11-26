import tensorflow as tf
import numpy as np
import os
from matplotlib import pyplot as plt
from openpyxl import Workbook, load_workbook

class SparsityAccuracyCallback(tf.keras.callbacks.Callback):
    def __init__(self, model,final_sparsity, accuracy_file, loss_file):
        super(SparsityAccuracyCallback, self).__init__()
        self.model = model  # 存储模型
        self.final_sparsity = final_sparsity  # 当前目标稀疏度
        self.accuracy_file = accuracy_file  # 存储 accuracy 的文件路径
        self.loss_file = loss_file  # 存储 loss 的文件路径
        self.sparsity = []  # 存储稀疏度
        self.val_accuracy = []  # 存储验证集准确度
        self.val_loss = []  # 存储验证集损失
        self.epoch_data = []  # 用于存储 epoch
        # 初始化 Excel 文件
        self.accuracy_wb, self.accuracy_sheet = self._load_or_create_workbook(self.accuracy_file)
        self.loss_wb, self.loss_sheet = self._load_or_create_workbook(self.loss_file)

    def _load_or_create_workbook(self, file_path):
        """加载或创建 Excel 文件"""
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
            workbook.save(file_path)
        # 创建或获取工作表
        sheet_name = f"Sparsity_{self.final_sparsity:.1f}"
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(["Epoch", "Value"])  # 添加表头
        else:
            sheet = workbook[sheet_name]
        return workbook, sheet

    def on_epoch_end(self, epoch, logs=None):
        sparsity = calculate_sparsity(self.model)  # 计算稀疏度
        accuracy = logs.get('val_accuracy', None)  # 获取验证集准确度
        loss = logs.get('val_loss', None)  # 获取验证集损失
        self.sparsity.append(sparsity)  # 记录当前稀疏度
        self.val_accuracy.append(accuracy)  # 记录当前验证集准确度
        self.val_loss.append(loss)  # 记录当前验证集损失
        self.epoch_data.append(epoch + 1)

        # 写入 accuracy 文件
        self.accuracy_sheet.append([epoch + 1, accuracy])
        self.accuracy_wb.save(self.accuracy_file)

        # 写入 loss 文件
        self.loss_sheet.append([epoch + 1, loss])
        self.loss_wb.save(self.loss_file)

        print(f"\nEpoch {epoch + 1}: Sparsity = {sparsity:.2f}%, Validation Accuracy = {accuracy:.4f}, Validation Loss = {loss:.4f}")

    def plot_results(self,save_dir=None, filename="sparsity_accuracy_results.png"):
        # 绘制稀疏度与验证准确度的关系图
        plt.figure(figsize=(12, 6))

        # 左图：稀疏度 vs 验证准确度
        plt.subplot(1, 2, 1)
        plt.plot(self.sparsity, self.val_accuracy, marker='o', label='Validation Accuracy')
        plt.xlabel('Sparsity (%)')
        plt.ylabel('Validation Accuracy')
        plt.title('Validation Accuracy vs Sparsity')
        plt.grid()
        plt.legend()

        # 右图：稀疏度 vs 验证损失
        plt.subplot(1, 2, 2)
        plt.plot(self.sparsity, self.val_loss, marker='o', label='Validation Loss')
        plt.xlabel('Sparsity (%)')
        plt.ylabel('Validation Loss')
        plt.title('Validation Loss vs Sparsity')
        plt.grid()
        plt.legend()

        plt.tight_layout()
        plt.show()

        if save_dir:
            os.makedirs(save_dir, exist_ok=True)  # 确保目录存在
            save_path = os.path.join(save_dir, filename)
            plt.savefig(save_path, bbox_inches='tight')  # 保存图表
            print(f"Results plot saved to {save_path}")
            plt.close()  # 关闭图表，避免影响后续绘图
        else:
            plt.show()

    def on_train_end(self, logs=None):
        """训练结束时保存文件，确保所有数据写入"""
        self.accuracy_wb.save(self.accuracy_file)
        self.loss_wb.save(self.loss_file)

def calculate_sparsity(model):
    total_weights = 0  # 权重总数
    zero_weights = 0  # 零权重的个数

    for layer in model.layers:
        # 获取每一层的权重
        weights = layer.get_weights()
        if len(weights) > 0:  # 如果该层有权重
            # 仅考虑可剪枝的权重
            for weight in weights:
                total_weights += weight.size  # 累加该层权重的总个数
                zero_weights += np.sum(weight == 0)  # 累加零权重的个数

    # 计算并返回稀疏度百分比
    if total_weights == 0:  # 避免除零错误
        return 0
    return (zero_weights / total_weights) * 100

